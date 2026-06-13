#!/usr/bin/env python3
"""
Offer Tracker — Multi-user web app with browser OAuth, notification prefs, and demo mode.
"""
import os, sys, threading, time, json, base64
from datetime import datetime
from pathlib import Path
from email.mime.text import MIMEText

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, render_template, jsonify, redirect, request, session
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import Flow
from src.config import logger, POLL_INTERVAL_MINUTES, BASE_DIR, TELEGRAM_BOT_TOKEN, TELEGRAM_BOT_USERNAME, CRON_SECRET, GMAIL_QUERY
from src.notifier import notify_single
import requests as http_requests

app = Flask(__name__)

@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store"
    return response

_secret_key_file = BASE_DIR / ".secret_key"
if _secret_key_file.exists():
    app.secret_key = _secret_key_file.read_text().strip()
else:
    app.secret_key = os.urandom(24).hex()
    _secret_key_file.write_text(app.secret_key)

SCOPES = [
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
    "https://www.googleapis.com/auth/gmail.modify",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/spreadsheets",
]

CREDENTIALS_DIR = BASE_DIR / "credentials"
CREDENTIALS_PATH = CREDENTIALS_DIR / "credentials.json"
PREFS_PATH = BASE_DIR / "user_prefs.json"
SHEET_HEADERS = ["Company Name", "Job Role", "Application Date", "Email Subject", "Sender Email", "Message ID", "Alert Sent", "Email Type", "Summary", "Location", "Salary", "Next Step", "Parser"]

last_run: dict[str, datetime | None] = {}
last_count: dict[str, int] = {}
last_error: dict[str, str] = {}
log_buffer: dict[str, list[str]] = {}
_prefs_lock = threading.Lock()


def log(email: str, msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    log_buffer.setdefault(email, []).append(f"[{ts}] {msg}")
    logger.info(f"[{email}] {msg}")


def normalize_email(email: str) -> str:
    local, _, domain = email.partition('@')
    domain = domain.lower()
    if domain in ('gmail.com', 'googlemail.com'):
        local = local.replace('.', '').lower()
    return f"{local}@{domain}"

def get_user_email() -> str | None:
    email = session.get("user_email")
    if not email:
        return None
    norm = normalize_email(email)
    if norm != email:
        prefs = load_prefs()
        if email in prefs and norm not in prefs:
            prefs[norm] = prefs.pop(email)
            save_prefs(prefs)
        old_token = get_token_path(email)
        new_token = get_token_path(norm)
        if old_token.exists() and not new_token.exists():
            old_token.rename(new_token)
        for d in (last_run, last_count, last_error, log_buffer):
            if email in d:
                d[norm] = d.pop(email)
        session["user_email"] = norm
    return norm


def load_prefs() -> dict:
    with _prefs_lock:
        if PREFS_PATH.exists():
            return json.loads(PREFS_PATH.read_text())
        return {}


def save_prefs(prefs: dict) -> None:
    with _prefs_lock:
        PREFS_PATH.write_text(json.dumps(prefs, indent=2))


def get_user_prefs(email: str) -> dict:
    prefs = load_prefs()
    if email in prefs:
        return prefs[email]
    for k, v in prefs.items():
        if normalize_email(k) == email:
            return v
    return {}


def set_user_pref(email: str, key: str, value: str) -> None:
    prefs = load_prefs()
    norm = normalize_email(email)
    for k in list(prefs.keys()):
        if k != norm and normalize_email(k) == norm:
            prefs[norm] = {**prefs.pop(k), **prefs.get(norm, {})}
    prefs.setdefault(norm, {})[key] = value
    save_prefs(prefs)


def get_token_path(email: str) -> Path:
    safe = base64.urlsafe_b64encode(email.encode()).decode().rstrip("=")
    return CREDENTIALS_DIR / f"token_{safe}.json"


def get_creds(email: str) -> Credentials | None:
    path = get_token_path(email)
    if not path.exists():
        for f in CREDENTIALS_DIR.glob("token_*.json"):
            stored = base64.urlsafe_b64decode(f.stem.replace("token_", "") + "==").decode()
            if normalize_email(stored) == normalize_email(email) and stored != email:
                f.rename(path)
                break
    if not path.exists():
        return None
    try:
        creds = Credentials.from_authorized_user_file(str(path), SCOPES)
    except Exception as e:
        logger.warning(f"Token scope mismatch for {email}: {e}")
        path.unlink(missing_ok=True)
        return None
    if creds.expired and creds.refresh_token:
        try:
            creds.refresh(Request())
            path.write_text(creds.to_json())
        except Exception as e:
            logger.warning(f"Token refresh failed for {email}: {e}")
            return None
    return creds


def get_gmail_service(email: str):
    return build("gmail", "v1", credentials=get_creds(email))


def get_sheets_service(email: str):
    return build("sheets", "v4", credentials=get_creds(email))


def _find_or_create_sheet(email: str) -> str:
    prefs = get_user_prefs(email)
    sid = prefs.get("sheet_id")
    if sid:
        return sid
    svc = build("sheets", "v4", credentials=get_creds(email))
    sh = svc.spreadsheets().create(body={"properties": {"title": "Job Application Tracker"}}).execute()
    sid = sh["spreadsheetId"]
    set_user_pref(email, "sheet_id", sid)
    log(email, f"Created sheet: {sid}")
    return sid


def ensure_sheet(email: str) -> str:
    sid = _find_or_create_sheet(email)
    sheets = build("sheets", "v4", credentials=get_creds(email))
    result = sheets.spreadsheets().values().get(spreadsheetId=sid, range="A1:M1").execute()
    existing = result.get("values", [[]])[0]
    if existing != SHEET_HEADERS[:len(existing)]:
        sheets.spreadsheets().values().update(
            spreadsheetId=sid, range="A1",
            valueInputOption="RAW",
            body={"values": [SHEET_HEADERS]}
        ).execute()
    return f"https://docs.google.com/spreadsheets/d/{sid}"


def format_sheet(email: str) -> bool:
    sid = _find_or_create_sheet(email)
    sheets = build("sheets", "v4", credentials=get_creds(email))

    header_bg = {"red": 0.16, "green": 0.25, "blue": 0.45}
    header_fg = {"red": 1, "green": 1, "blue": 1}
    accent = {"red": 0.20, "green": 0.34, "blue": 0.57}
    light = {"red": 0.96, "green": 0.97, "blue": 0.99}
    white = {"red": 1, "green": 1, "blue": 1}

    requests = []

    try:
        meta = sheets.spreadsheets().get(spreadsheetId=sid, fields="sheets.bandedRanges").execute()
        existing = meta.get("sheets", [{}])[0].get("bandedRanges", [])
        for br in existing:
            rid = br.get("bandedRangeId")
            if rid:
                requests.append({"deleteBanding": {"bandedRangeId": rid}})
    except Exception:
        pass

    requests.append({
        "addBanding": {
            "bandedRange": {
                "range": {"sheetId": 0, "startRowIndex": 1},
                "rowProperties": {
                    "headerColor": header_bg,
                    "firstBandColor": light,
                    "secondBandColor": white,
                }
            }
        }
    })

    requests.append({
        "repeatCell": {
            "range": {"sheetId": 0, "startRowIndex": 0, "endRowIndex": 1},
            "cell": {
                "userEnteredFormat": {
                    "backgroundColor": header_bg,
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "textFormat": {
                        "bold": True,
                        "foregroundColor": header_fg,
                        "fontSize": 11,
                    },
                }
            },
            "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat.bold,textFormat.foregroundColor,textFormat.fontSize)"
        }
    })

    requests.append({
        "updateBorders": {
            "range": {"sheetId": 0, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 13},
            "bottom": {"style": "SOLID", "width": 2, "color": accent},
        }
    })

    requests.append({
        "autoResizeDimensions": {
            "dimensions": {"sheetId": 0, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 13}
        }
    })

    requests.append({
        "updateSheetProperties": {
            "properties": {"sheetId": 0, "gridProperties": {"frozenRowCount": 1}},
            "fields": "gridProperties.frozenRowCount"
        }
    })

    FONT_SIZE = 10
    wrap_fields = "userEnteredFormat(textFormat.fontSize,wrapStrategy)"
    align_fields = "userEnteredFormat(horizontalAlignment,verticalAlignment)"
    all_fmt_fields = f"{wrap_fields},{align_fields}"

    requests.append({
        "repeatCell": {
            "range": {"sheetId": 0, "startRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 13},
            "cell": {
                "userEnteredFormat": {
                    "textFormat": {"fontSize": FONT_SIZE},
                    "wrapStrategy": "WRAP",
                    "verticalAlignment": "MIDDLE",
                }
            },
            "fields": wrap_fields
        }
    })

    date_cols = [2]
    for ci in date_cols:
        requests.append({
            "repeatCell": {
                "range": {"sheetId": 0, "startRowIndex": 1, "startColumnIndex": ci, "endColumnIndex": ci + 1},
                "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER"}},
                "fields": "userEnteredFormat.horizontalAlignment"
            }
        })

    try:
        sheets.spreadsheets().batchUpdate(spreadsheetId=sid, body={"requests": requests}).execute()
        log(email, "Sheet formatted beautifully")
        return True
    except Exception as e:
        log(email, f"Format failed: {e}")
        return False


def send_test_email(email: str) -> None:
    service = get_gmail_service(email)
    msg = MIMEText(
        "Dear Candidate,\n\nThank you for your application for the Software Engineer Intern position at Google.\n"
        "We have received your application and will review it shortly.\n\nBest regards,\nGoogle Hiring Team"
    )
    msg["To"] = msg["From"] = email
    msg["Subject"] = "Application Received: Software Engineer Intern at Google"
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    log(email, "Test email sent")


def run_poll(email: str) -> None:
    try:
        creds = get_creds(email)
        if not creds:
            raise RuntimeError("No credentials")

        sheet_url = ensure_sheet(email)
        gmail = build("gmail", "v1", credentials=creds)
        sheets = build("sheets", "v4", credentials=creds)

        results = gmail.users().messages().list(
            userId="me",
            q=f"({GMAIL_QUERY}) is:unread",
            maxResults=20,
        ).execute()

        msgs = results.get("messages", [])
        if not msgs:
            log(email, "No matching emails found")
            last_run[email] = datetime.now()
            last_count[email] = 0
            return

        sid = _find_or_create_sheet(email)
        known_ids = set()
        existing = sheets.spreadsheets().values().get(spreadsheetId=sid, range="F:F").execute()
        if existing.get("values"):
            known_ids = {r[0] for r in existing["values"][1:] if r}

        existing_all = sheets.spreadsheets().values().get(spreadsheetId=sid, range="A:M").execute()
        existing_rows = existing_all.get("values", [])[1:] if existing_all.get("values") else []
        company_role_map = {}
        for ri, row in enumerate(existing_rows):
            if len(row) >= 13:
                key = (row[0].strip().lower(), row[1].strip().lower())
                company_role_map[key] = (ri + 2, row[7])

        STATUS_PRIORITY = {
            "rejection": 1, "other": 2, "application_received": 3,
            "assessment": 4, "phone_screen": 5, "interview_invitation": 6,
            "technical_interview": 7, "offer_letter": 8,
        }

        count = 0

        for m in msgs:
            try:
                full = gmail.users().messages().get(userId="me", id=m["id"], format="full").execute()
                headers = {h["name"]: h["value"] for h in full["payload"]["headers"]}
                msg_id = headers.get("Message-ID", "")
                if msg_id in known_ids:
                    continue

                from src.parser import parse_email
                app = parse_email(full)
                if not app:
                    gmail.users().messages().modify(userId="me", id=m["id"], body={"removeLabelIds": ["UNREAD"]}).execute()
                    continue

                match_key = (app.company_name.strip().lower(), app.job_role.strip().lower())
                if match_key in company_role_map:
                    row_num, old_type_label = company_role_map[match_key]
                    old_type = next((k for k, v in EMAIL_TYPE_LABEL.items() if v == old_type_label), "other")
                    new_priority = STATUS_PRIORITY.get(app.email_type, 0)
                    old_priority = STATUS_PRIORITY.get(old_type, 0)
                    if new_priority > old_priority:
                        sheet_row = app.to_sheet_row()
                        body = {"values": [sheet_row]}
                        sheets.spreadsheets().values().update(
                            spreadsheetId=sid,
                            range=f"A{row_num}:M{row_num}",
                            valueInputOption="RAW",
                            body=body,
                        ).execute()
                        logger.info(f"Updated row {row_num}: {app.company_name} - {app.job_role} ({old_type_label} -> {app.email_type})")
                    else:
                        logger.debug(f"Skipped update for {app.company_name} - {app.job_role}: {app.email_type} <= {old_type_label}")
                        gmail.users().messages().modify(userId="me", id=m["id"], body={"removeLabelIds": ["UNREAD"]}).execute()
                        continue
                else:
                    sheets.spreadsheets().values().append(
                        spreadsheetId=sid, range="A:M",
                        valueInputOption="RAW",
                        insertDataOption="INSERT_ROWS",
                        body={"values": [app.to_sheet_row()]}
                    ).execute()
                    known_ids.add(msg_id)

                from src.notifier import notify_single
                prefs = get_user_prefs(email)
                notify_single(prefs.get("notification_channel"), prefs, app.to_alert_text(), email)

                gmail.users().messages().modify(userId="me", id=m["id"], body={"removeLabelIds": ["UNREAD"]}).execute()
                count += 1
                log(email, f"[{app.parser}] Logged: {app.company_name} - {app.job_role}")
            except Exception as e:
                log(email, f"Failed to process message: {e}")
                import traceback
                logger.error(f"Traceback for email {m['id']}:\n{traceback.format_exc()}")
                try:
                    gmail.users().messages().modify(userId="me", id=m["id"], body={"removeLabelIds": ["UNREAD"]}).execute()
                except Exception:
                    pass

        last_run[email] = datetime.now()
        last_count[email] = count
        log(email, f"Checked: {count} new entries")
    except Exception as e:
        last_error[email] = str(e)
        log(email, f"Poll error: {e}")


def scheduler_loop() -> None:
    time.sleep(10)
    while True:
        for email in load_prefs():
            if get_creds(email):
                run_poll(email)
        time.sleep(POLL_INTERVAL_MINUTES * 60)


# ── Shared helpers ─────────────────────────────────────────

def _get_dashboard_context(email: str) -> dict:
    prefs = get_user_prefs(email)
    sheet_url = rows = sheet_error = ""
    try:
        sheet_url = ensure_sheet(email)
        svc = get_sheets_service(email)
        sid = _find_or_create_sheet(email)
        data = svc.spreadsheets().values().get(spreadsheetId=sid, range="A:M").execute()
        vals = data.get("values", [])
        if vals:
            rows = vals[1:][-20:]
    except Exception as e:
        sheet_error = str(e)
    return dict(
        authed=True, email=email, prefs=prefs,
        prefs_complete="notification_channel" in prefs,
        sheet_url=sheet_url, rows=rows,
        headers=SHEET_HEADERS[:5] + ["Stage", "Summary", "Location", "Salary", "Next Step", "Parser"],
        last_run=last_run.get(email), last_count=last_count.get(email, 0),
        last_error=last_error.get(email, ""), sheet_error=sheet_error,
        interval=POLL_INTERVAL_MINUTES,
        logs=log_buffer.get(email, [])[-30:],
        telegram_bot_username=TELEGRAM_BOT_USERNAME,
        has_telegram_bot=bool(TELEGRAM_BOT_TOKEN),
        scheduler_alive=_scheduler_alive["running"],
        alert_message="",
    )


# ── Routes ────────────────────────────────────────────────

@app.route("/")
def index():
    email = get_user_email()
    if not email or not get_creds(email):
        setup_needed = not CREDENTIALS_PATH.exists()
        return render_template("index.html", authed=False, needs_creds=setup_needed, redirect_uri=redirect_uri())
    return render_template("index.html", **_get_dashboard_context(email))


@app.route("/auth")
def auth():
    if not CREDENTIALS_PATH.exists():
        return redirect("/")
    flow = Flow.from_client_secrets_file(str(CREDENTIALS_PATH), scopes=SCOPES, redirect_uri=redirect_uri())
    flow.autogenerate_code_verifier = True
    auth_url, state = flow.authorization_url(access_type="offline", prompt="consent")
    session["oauth_state"] = state
    session["code_verifier"] = flow.code_verifier
    return redirect(auth_url)


@app.route("/callback")
def callback():
    if "oauth_state" not in session:
        return redirect("/")
    try:
        flow = Flow.from_client_secrets_file(str(CREDENTIALS_PATH), scopes=SCOPES, redirect_uri=redirect_uri())
        flow.code_verifier = session.pop("code_verifier", None)
        flow.fetch_token(authorization_response=request.url, state=session.pop("oauth_state"))

        userinfo = build("oauth2", "v2", credentials=flow.credentials).userinfo().get().execute()
        email = userinfo.get("email") or userinfo.get("name", "")

        CREDENTIALS_DIR.mkdir(parents=True, exist_ok=True)
        get_token_path(email).write_text(flow.credentials.to_json())
        session["user_email"] = normalize_email(email)
        log(email, "Gmail connected")
        return redirect("/")
    except Exception as e:
        logger.error(f"OAuth callback failed: {e}")
        return render_template("index.html", authed=False, needs_creds=False, redirect_uri=redirect_uri(), oauth_error=str(e))


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


@app.route("/save-prefs", methods=["POST"])
def save_prefs_route():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401

    channel = request.form.get("channel", "none")
    if channel not in ("whatsapp", "slack", "telegram", "discord", "whatsapp_cloud", "twilio_whatsapp", "pushover", "none"):
        return jsonify({"error": "Invalid channel"}), 400

    if channel == "none":
        set_user_pref(email, "notification_channel", "none")
        log(email, "Notifications disabled")
        return jsonify({"status": "ok"})

    set_user_pref(email, "notification_channel", channel)

    if channel == "whatsapp":
        phone = request.form.get("whatsapp_phone", "").strip()
        if not phone:
            return jsonify({"error": "Phone number required"}), 400
        set_user_pref(email, "whatsapp_phone", phone)
        apikey = request.form.get("whatsapp_apikey", "").strip()
        if apikey:
            set_user_pref(email, "whatsapp_apikey", apikey)
    elif channel == "pushover":
        user_key = request.form.get("pushover_user_key", "").strip()
        if not user_key:
            return jsonify({"error": "Pushover User Key required"}), 400
        set_user_pref(email, "pushover_user_key", user_key)
    elif channel == "slack":
        webhook_url = request.form.get("slack_webhook_url", "").strip()
        if not webhook_url:
            return jsonify({"error": "Webhook URL required"}), 400
        if not webhook_url.startswith("https://hooks.slack.com/"):
            return jsonify({"error": "Invalid Slack webhook URL"}), 400
        set_user_pref(email, "slack_webhook_url", webhook_url)
    elif channel == "discord":
        webhook_url = request.form.get("discord_webhook_url", "").strip()
        if not webhook_url:
            return jsonify({"error": "Webhook URL required"}), 400
        if not webhook_url.startswith("https://discord.com/api/webhooks/"):
            return jsonify({"error": "Invalid Discord webhook URL"}), 400
        set_user_pref(email, "discord_webhook_url", webhook_url)
    elif channel == "telegram":
        username = request.form.get("telegram_username", "").strip().lstrip("@")
        if not username:
            return jsonify({"error": "Telegram username required"}), 400
        prefs = load_prefs()
        saved_username = prefs.get(email, {}).get("telegram_username", "")
        saved_chat_id = prefs.get(email, {}).get("telegram_chat_id", "")
        set_user_pref(email, "telegram_username", username)
        if username != saved_username:
            set_user_pref(email, "telegram_chat_id", "")
    elif channel == "whatsapp_cloud":
        phone = request.form.get("whatsapp_cloud_phone", "").strip()
        if not phone:
            return jsonify({"error": "Phone number required"}), 400
        set_user_pref(email, "whatsapp_cloud_phone", phone)
    elif channel == "twilio_whatsapp":
        phone = request.form.get("twilio_whatsapp_phone", "").strip()
        if not phone:
            return jsonify({"error": "Phone number required"}), 400
        set_user_pref(email, "twilio_whatsapp_phone", phone)

    log(email, f"Notification channel set: {channel}")
    return jsonify({"status": "ok"})


@app.route("/trigger")
def trigger():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    try:
        run_poll(email)
        _alert = f"Poll complete — {last_count.get(email, 0)} new entries"
    except Exception as e:
        _alert = f"Poll failed: {e}"
        last_error[email] = str(e)
        log(email, _alert)
    ctx = _get_dashboard_context(email)
    ctx["alert_message"] = _alert
    return render_template("_dashboard.html", **ctx)


@app.route("/verify-telegram")
def verify_telegram():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    if not TELEGRAM_BOT_TOKEN:
        return jsonify({"error": "Telegram bot not configured by admin"}), 400

    prefs = get_user_prefs(email)
    username = prefs.get("telegram_username", "")
    if not username:
        return jsonify({"error": "No Telegram username saved"}), 400

    try:
        resp = http_requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates",
            json={"limit": 10, "timeout": 0},
            timeout=10,
        )
        data = resp.json()
        if not data.get("ok"):
            return jsonify({"error": f"Telegram API error: {data.get('description', 'unknown')}"}), 500

        for update in data.get("result", []):
            msg = update.get("message", {}) or update.get("channel_post", {})
            msg_user = msg.get("from", {})
            msg_username = (msg_user.get("username") or "").lower()
            if msg_username == username.lower():
                chat_id = str(msg["chat"]["id"])
                set_user_pref(email, "telegram_chat_id", chat_id)
                log(email, f"Telegram verified — chat_id: {chat_id}")
                return jsonify({"status": "ok", "chat_id": chat_id})

        log(email, f"No message found from @{username} — waiting for them to DM the bot")
        return jsonify({"status": "waiting", "message": f"Waiting for @{username} to message the bot..."})
    except Exception as e:
        log(email, f"Telegram verify error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/change-channel")
def change_channel():
    email = get_user_email()
    if not email:
        return redirect("/")
    prefs = load_prefs()
    if email in prefs and "notification_channel" in prefs[email]:
        del prefs[email]["notification_channel"]
        save_prefs(prefs)
    log(email, "Channel reset — choose a new one")
    return redirect("/")


@app.route("/dedup-sheet")
def dedup_sheet_route():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    try:
        sid = _find_or_create_sheet(email)
        sheets = build("sheets", "v4", credentials=get_creds(email))
        result = sheets.spreadsheets().values().get(spreadsheetId=sid, range="A:M").execute()
        rows = result.get("values", [])
        if not rows:
            return jsonify({"status": "ok", "removed": 0})
        seen = {}
        keep = [rows[0]]
        dup_count = 0
        for row in rows[1:]:
            mid = row[5] if len(row) > 5 else ""
            if mid and mid in seen:
                dup_count += 1
                continue
            if mid:
                seen[mid] = True
            keep.append(row)
        if dup_count > 0:
            sheets.spreadsheets().values().clear(spreadsheetId=sid, range="A2:M").execute()
            if len(keep) > 1:
                sheets.spreadsheets().values().update(
                    spreadsheetId=sid, range="A2",
                    valueInputOption="RAW",
                    body={"values": keep[1:]}
                ).execute()
        log(email, f"Dedup: removed {dup_count} duplicates, kept {len(keep)-1} entries")
        return jsonify({"status": "ok", "removed": dup_count, "kept": len(keep) - 1})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/format-sheet")
def format_sheet_route():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    ok = format_sheet(email)
    return jsonify({"status": "ok" if ok else "error"})

@app.route("/send-test-email")
def send_test_email_route():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    try:
        send_test_email(email)
        log(email, "Test email sent")
        _alert = "Test email sent to your Gmail inbox!"
    except Exception as e:
        _alert = f"Test email failed: {e}"
        log(email, _alert)
    ctx = _get_dashboard_context(email)
    ctx["alert_message"] = _alert
    return render_template("_dashboard.html", **ctx)


@app.route("/status")
def status():
    email = get_user_email()
    if not email:
        return jsonify({"authed": False})
    p = get_user_prefs(email)
    return jsonify({
        "authed": True, "email": email,
        "prefs_complete": "notification_channel" in p,
        "channel": p.get("notification_channel"),
        "last_run": last_run.get(email).isoformat() if last_run.get(email) else None,
        "last_count": last_count.get(email, 0),
        "last_error": last_error.get(email, ""),
    })


@app.route("/logs")
def logs():
    email = get_user_email()
    return jsonify(log_buffer.get(email, [])[-50:] if email else [])


_scheduler_alive = {"running": True, "started": datetime.now().isoformat()}

@app.route("/cron/<secret>")
def cron_trigger(secret: str):
    if secret != CRON_SECRET:
        return jsonify({"error": "invalid secret"}), 403
    if not CRON_SECRET:
        return jsonify({"error": "cron not configured"}), 400
    emails = [e for e, p in load_prefs().items() if get_creds(e)]
    results = {}
    for email in emails:
        try:
            run_poll(email)
            results[email] = {"status": "ok", "count": last_count.get(email, 0)}
        except Exception as e:
            results[email] = {"status": "error", "error": str(e)}
    return jsonify({"status": "ok", "users": len(emails), "results": results})


@app.route("/automation-status")
def automation_status():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    return jsonify({
        "scheduler_alive": True,
        "scheduler_started": _scheduler_alive["started"],
        "last_run": last_run.get(email).isoformat() if last_run.get(email) else None,
        "poll_interval_minutes": POLL_INTERVAL_MINUTES,
    })


@app.route("/test-notification", methods=["POST"])
def test_notification():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    prefs = get_user_prefs(email)
    channel = prefs.get("notification_channel")
    if not channel or channel == "none":
        msg = "No notification channel configured"
        log(email, msg)
        return jsonify({"status": "error", "error": msg}), 400
    msg = "Test notification — your Offer Tracker is working!"
    try:
        notify_single(channel, prefs, msg, email)
        _alert = f"Test {channel} notification sent!"
        log(email, _alert)
    except Exception as e:
        _alert = f"Test notification failed: {e}"
        log(email, _alert)
    if request.headers.get("HX-Request"):
        ctx = _get_dashboard_context(email)
        ctx["alert_message"] = _alert
        return render_template("_dashboard.html", **ctx)
    return jsonify({"status": _alert.startswith("Test") and "ok" or "error", "message": _alert})


@app.route("/save-whatsapp-apikey", methods=["POST"])
def save_whatsapp_apikey():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    apikey = request.form.get("apikey", "").strip()
    if not apikey:
        return jsonify({"error": "API key required"}), 400
    set_user_pref(email, "whatsapp_apikey", apikey)
    log(email, "WhatsApp API key saved")
    return jsonify({"status": "ok"})


@app.route("/save-pushover-key", methods=["POST"])
def save_pushover_key():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    key = request.form.get("pushover_user_key", "").strip()
    if not key:
        return jsonify({"error": "User Key required"}), 400
    set_user_pref(email, "pushover_user_key", key)
    log(email, "Pushover user key saved")
    return jsonify({"status": "ok"})


@app.route("/download-contacts")
def download_contacts():
    vcf = "BEGIN:VCARD\nVERSION:3.0\nFN:CallMeBot\nTEL;TYPE=CELL:+34644599043\nEND:VCARD\n"
    from flask import make_response
    resp = make_response(vcf)
    resp.headers["Content-Type"] = "text/vcard"
    resp.headers["Content-Disposition"] = 'attachment; filename="callmebot_contacts.vcf"'
    return resp


@app.route("/export-xlsx")
def export_xlsx():
    email = get_user_email()
    if not email:
        return jsonify({"error": "Not authenticated"}), 401
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        sid = _find_or_create_sheet(email)
        sheets = get_sheets_service(email)
        result = sheets.spreadsheets().values().get(
            spreadsheetId=sid, range="A:M"
        ).execute()
        values = result.get("values", [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"

        header_fill = PatternFill(start_color="2A4073", end_color="2A4073", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(bottom=Side(style="thin", color="2A4073"))

        for col_idx, header in enumerate(SHEET_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        light_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
        for row_idx, row in enumerate(values[1:], 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if row_idx % 2 == 0:
                    cell.fill = light_fill
                if col_idx == 3:
                    cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

        ws.freeze_panes = "A2"

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import send_file
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"job_applications_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    except Exception as e:
        logger.error(f"XLSX export failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/upload-credentials", methods=["POST"])
def upload_credentials():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".json"):
        return jsonify({"error": "Must be .json"}), 400
    CREDENTIALS_DIR.mkdir(parents=True, exist_ok=True)
    f.save(str(CREDENTIALS_PATH))
    return jsonify({"status": "ok"})


def redirect_uri() -> str:
    scheme = request.headers.get("X-Forwarded-Proto", request.scheme)
    return f"{scheme}://{request.host}/callback"


threading.Thread(target=scheduler_loop, daemon=True).start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    print(f"Starting on port {port}")
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
